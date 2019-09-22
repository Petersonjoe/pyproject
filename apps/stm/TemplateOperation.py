# -*- coding: utf-8 -*-
#############################################
# @Author: jlei1
# @Date:   2018-07-06
# @Last Modified By:   jlei1
# @Last Modified Time: 2018-11-07
#############################################

from distutils.log import warn as printf
import os, sys
from time import localtime, strftime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

if '../' not in sys.path:
	sys.path.insert(0,'../')
from utils.ReadConfig import DecodeConfig
from utils.GlobalVars import PROJ_ROOT_DIR
from .DataMapper import *

TMPL_CONFIG = {
	'TMPL_SRC_DIR': None,
	'TMPL_DST_DIR': None,
	'TMPL_SRC_FILE': None,
	'TMPL_DST_FILE': None,
	'TMPL_SHT_NAME': None,
	'TMPL_BGN_COL': None,
	'TMPL_END_COL': None,
	'TMPL_HD_ROW': None,
	'TMPL_FT_ROW': None
}

# read config file
config_path = PROJ_ROOT_DIR + '/conf/'
config_name = 'global'
config = DecodeConfig()
tmpl_settings = config.getConfig(filepath=config_path, filename=config_name)	

# set template path info
TMPL_CONFIG['TMPL_SRC_DIR'] = PROJ_ROOT_DIR + tmpl_settings['stm_template']['tmpl_src_dir']
TMPL_CONFIG['TMPL_DST_DIR'] = PROJ_ROOT_DIR + tmpl_settings['stm_template']['tmpl_dst_dir']
TMPL_CONFIG['TMPL_SRC_FILE'] = tmpl_settings['stm_template']['tmpl_src_file']
TMPL_CONFIG['TMPL_DST_FILE'] = tmpl_settings['stm_template']['tmpl_dst_file']
src_path = TMPL_CONFIG['TMPL_SRC_DIR'] + TMPL_CONFIG['TMPL_SRC_FILE']
dst_path = TMPL_CONFIG['TMPL_DST_DIR'] + TMPL_CONFIG['TMPL_DST_FILE']
output_dir = PROJ_ROOT_DIR + '/output/stm/tmp/'
output_name = '_out_' + strftime("%Y_%m_%d_%H_%M_%S", localtime()) + '.xlsx'

# set template data
TMPL_CONFIG['TMPL_SHT_NAME'] = tmpl_settings['stm_template']['tmpl_sht_name']
TMPL_CONFIG['TMPL_BGN_COL'] = tmpl_settings['stm_template']['tmpl_bgn_col']
TMPL_CONFIG['TMPL_END_COL'] = tmpl_settings['stm_template']['tmpl_end_col']
TMPL_CONFIG['TMPL_HD_ROW'] = tmpl_settings['stm_template']['tmpl_hd_row']
TMPL_CONFIG['TMPL_FT_ROW'] = tmpl_settings['stm_template']['tmpl_ft_row']
NUM_OF_ADT = int(tmpl_settings['stm_template']['tmpl_adt_num'])

# set template format
MRG_COL_IX = [int(i) for i in tmpl_settings['stm_template_format']['merged_col_ix'].split(',')]
HDR_COLOR = tmpl_settings['stm_template_format']['header_color']
HDR_FONT = tmpl_settings['stm_template_format']['header_font']
HDR_FONT_SIZE = tmpl_settings['stm_template_format']['header_font_size']
HDR_FONT_BOLD = True if tmpl_settings['stm_template_format']['header_font_bold'] == 'Y' else False
DAT_COLOR = tmpl_settings['stm_template_format']['data_color']
DAT_FONT = tmpl_settings['stm_template_format']['data_font']
DAT_FONT_SIZE = tmpl_settings['stm_template_format']['data_font_size']
DAT_FONT_BOLD = True if tmpl_settings['stm_template_format']['data_font_bold'] == 'Y' else False
DAT_COL_BOLD_IX = [int(i) for i in tmpl_settings['stm_template_format']['data_col_bold_ix'].split(',')]

def setTemplate(replaceOld: bool = False) -> None:
	"""
	Set which template stm used. All file configuration can be found
	in global config file.
	:param1 replaceOld, default is False, if True, will replace current
	        template with the new one in configuration path
	"""
	# set the path of config file, read stm settings 
	if replaceOld:
		if os.path.isfile(dst_path):
			os.remove(dst_path)

	wb = load_workbook(filename = src_path)
	sheet_ranges = wb[TMPL_CONFIG['TMPL_SHT_NAME']]
	wb.template = True
	wb.save(dst_path)

	return

def mergeCellsWithStyle(
		ws: openpyxl.Workbook = None, 
		cell_range: openpyxl.Workbook = None, 
		border: Border = Border(), 
		fill: bool = None, 
		font: bool = None, 
		alignment: bool = None
	) -> None:
	"""
	Apply styles to a range of cells as if they were a single cell.
	Copy from openpyxl docs: http://openpyxl.readthedocs.io/en/stable/styles.html#cell-styles-and-named-styles
	
	:param1 ws:  Excel worksheet instance
	:param2 range: An excel range to style (e.g. A1:F20)
	:param3 border: An openpyxl Border
	:param4 fill: An openpyxl PatternFill or GradientFill
	:param5 font: An openpyxl Font object
	"""

	top = Border(top=border.top)
	left = Border(left=border.left)
	right = Border(right=border.right)
	bottom = Border(bottom=border.bottom)

	first_cell = ws[cell_range.split(":")[0]]
	if alignment:
		ws.merge_cells(cell_range)
		first_cell.alignment = alignment

	rows = ws[cell_range]
	if font:
		first_cell.font = font

	for cell in rows[0]:
		cell.border = cell.border + top
	for cell in rows[-1]:
		cell.border = cell.border + bottom

	for row in rows:
		l = row[0]
		r = row[-1]
		l.border = l.border + left
		r.border = r.border + right
		if fill:
			for c in row:
				c.fill = fill

class TemplateOperations(object):
	"""Define processes for a excel template to automate a specific task.
	   New automation task can inherit this class and overwrite specific method which is needed.
	"""

	def __init__(self):
		super(TemplateOperations, self).__init__()
		# self.arg = arg

	def __checkPII__(self):
		pass

	def __checkAbbr__(self, col_name: str = None) -> str:
		'''__checkAbbr__ will return a abbreviated string according to
		   the rules given by business requirements.
           
		   :param col_name, the initial column name from ORACLE db
		   See the configuration file for changing any abbreviation mapping.
		'''
		# below assert declarations cannot work with exec() method
		# assert col_name is None, 'Error: Cannot abbreviate None type!'
		# assert isinstance(col_name, str), 'Error: Parameter must be STRING!'

		_col_name = None
		if col_name is not None:
			for key, value in ABBR_DICT.items():
				col_name = col_name.replace(key.upper(), value)
			_col_name = col_name
		
		return _col_name

	def __checkCharSet__(self, col_type: str = None) -> str:
		'''__checkCharSet__ will return a string value - 'LATIN' or '',
		   decided by the original column type.

		   :param col_type, a string to tag a column's data type,
		   oringinally from ORACLE db.
		'''

		if (col_type[0:7].upper() == 'VARCHAR') \
		   or (col_type[0:4].upper() == 'CHAR'):
			return 'LATIN'
		else:
			return ''
		
	def __mapTdDataType__(self, in_type: str = None) -> str:
		'''__mapTdDataType__ will return the corresponding teradata
		   data type for a ORACLE data type, the return value is a 
		   string `VARCHAR`, `DECIMAL` or so.
			
		   :param in_type, the data type from ORACLE system.
		'''
		_in_type = in_type.upper()

		if (_in_type[0:8] == 'VARCHAR2'):
			_out_type = ORA_2_TD[_in_type[0:8]] + _in_type[8:len(_in_type)]
		elif (_in_type[0:6] == 'NUMBER'):
		   _out_type = ORA_2_TD[_in_type[0:6]] + _in_type[6:len(_in_type)]
		else:
			_out_type = ORA_2_TD[_in_type]
		
		return _out_type

	def __mapSpDataType__(self, in_type: str = None) -> str:
		'''__mapSpDataType__ will return the corresponding spark
		   data type for a ORACLE data type, the return value is 
		   a string `STRING`, `DECIMAL` or so.
			
		   :param in_type, the data type from ORACLE system.
		'''
		_in_type = in_type.upper()

		if (_in_type[0:8] == 'VARCHAR2'):
			_out_type = ORA_2_SP[_in_type[0:8]]
		elif (_in_type[0:6] == 'NUMBER'):
			_out_type = ORA_2_SP[_in_type[0:6]] + _in_type[6:len(_in_type)]
		else:
			_out_type = ORA_2_SP[_in_type]

		return _out_type

	def __formatDate__(self, in_type: str = None) -> str:
		if in_type.upper() in DATE_FORMAT.keys():
			return DATE_FORMAT[in_type.upper()]
		else:
			return None

	def __footerHandler__(self, 
			sa: str = None,
			f_data: list = None, 
			r_counter: int = None,
			r_num: int = None,
			c_num: int = None,
			t_name: str = None,
			xl_sheet: openpyxl.Workbook = None
		) -> None:
		'''This method is to handle the template footer data, append them to the tail of body data.

		   :param1 f_data, the data list read from the template file, element in this list is xlCell
		   :param2 r_counter, a flag to identify the current data row in xl_sheet
		   :param3 r_num, the number of rows in f_data
		   :param4 c_num, the number of columns in template
		   :param5 xl_sheet, the Workbook object that will be return
		'''
		assert (f_data is not None) \
		   and (r_counter is not None) \
		   and (r_num is not None) \
		   and (c_num is not None) \
		   and (t_name is not None) \
		   and (xl_sheet is not None), \
		   '''\n\tError: All params cannot be None! Check the template footers!
		   '''
		t_name_ix = [key-1 for key, value in TBL_ROW_IX.items() if 'tbl_name' in value.lower()] 

		tbl_name = t_name

		for i in range(r_num):
			r_counter += 1
			if (i % NUM_OF_ADT) < (NUM_OF_ADT / 2):
				for j in range(c_num):
					if (f_data[i][j].value is None) and (j not in t_name_ix):
						continue
					elif j in t_name_ix:
						# _ = xl_sheet.cell(row=r_counter, column=j+1, value="{0}".format(t_name))
						gdict = globals()
						ldict = locals()
						exec('xl_sheet.cell(row=r_counter, column=j+1, value={0})'.format(TBL_ROW_IX[j+1]),gdict,ldict) 
					else:
						_ = xl_sheet.cell(row=r_counter, column=j+1, value="{0}".format(f_data[i][j].value))
			else:
				for j in range(c_num):
					if f_data[i][j].value is None:
						continue
					else:
						_ = xl_sheet.cell(row=r_counter, column=j+1, value="{0}".format(f_data[i][j].value))
		
		return r_counter

	def __xlFormater__(self,
			xl_sheet: openpyxl.Workbook = None,
			hd_row_max: int = None,
			xls_max_rnum: int = None,
			xls_max_cnum: int = None,
			mrg_ix: list = MRG_COL_IX,
			col_bld_ix: list = DAT_COL_BOLD_IX,
			hd_color: str = HDR_COLOR,
			hd_fnt: str = HDR_FONT,
			hd_fnt_sz: int = HDR_FONT_SIZE,
			is_hd_bld: bool = HDR_FONT_BOLD,
			dat_color: str = DAT_COLOR,
			dat_fnt: str = DAT_FONT,
			dat_fnt_sz: int = DAT_FONT_SIZE,
			is_dat_bld: bool = DAT_FONT_BOLD,
			**kwagrs
		):
		'''Render styles of the stm result based on the template's format, 
		   need to read format from config file.
		   
		   :param xl_sheet,     the excel sheet to be rendered
		   :param hd_row_max,   the max row of the stm header
		   :param xls_max_rnum, the max row of the stm data
		   :param xls_max_cnum, the max column of the stm data
		   :param mrg_ix,       a list stored the column indexes to be cell-merged
		   :param col_bld_ix,   a list stored the column indexes are bold styled
		   :param hd_color,     the hex value of the header's background color
		   :param hd_fnt,       the font name of the header
		   :param hd_fnt_sz,    the font size of the header
		   :param is_hd_bld,    the flag for bolding header
		   :param dat_color,    the hex value of the data's background color
		   :param dat_fnt,      the font name of data rows
		   :param dat_fnt_sz,   the font size of data rows
		   :param is_dat_bld,   the flag for bolding data rows 
		'''
		assert (xl_sheet is not None) \
		   and (hd_row_max is not None) \
		   and (xls_max_rnum is not None) \
		   and (xls_max_cnum is not None), \
		   '''\n\tError: Params cannot be None! 
		      \n\tCheck parameters "xl_sheet", "hd_row_max", "xls_max_rnum" or "xls_max_cnum"!
		   '''
		from openpyxl.styles import NamedStyle
		    
		_count = 0
		# pre-define border type
		# in this pylib, 000000 is black, FFFFFF is white
		thin = Side(border_style="thin", color="000000")
		thick = Side(border_style="thick", color="000000")

		# handling header's style
		header_font = Font(name=hd_fnt,size=hd_fnt_sz,color='000000',bold=is_hd_bld,italic=False,vertAlign=None,underline='none',strike=False)
		header_fill = PatternFill(fill_type="solid",start_color=hd_color,end_color=hd_color)
		header_align = Alignment(horizontal='left',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
		header_border = Border(top=thick,left=thick,right=thick,bottom=thick)

		for i in range(hd_row_max):
			_count += 1
			for j in range(xls_max_cnum):
				xl_cell = xl_sheet.cell(i+1, j+1)
				xl_cell.font = header_font
				xl_cell.fill = header_fill
				xl_cell.border = header_border
				xl_cell.alignment = header_align

		# handling data's style
		data_font = Font(name=dat_fnt,size=dat_fnt_sz,color='000000',bold=is_dat_bld,italic=False,vertAlign=None,underline='none',strike=False)
		data_fill = PatternFill(fill_type="solid",start_color=dat_color,end_color=dat_color)
		data_align = Alignment(horizontal='left',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
		data_border = Border(top=thin,left=thin,right=thin,bottom=thin)

		for i in range(xls_max_rnum):
			if i+1 > _count:
				_count += 1
				for j in range(xls_max_cnum):
					xl_cell = xl_sheet.cell(i+1, j+1)
					xl_cell.font = data_font
					xl_cell.fill = data_fill
					xl_cell.border = data_border
					xl_cell.alignment = data_align

		# misc style settings
		# merge cells
		for ix in mrg_ix:
			xl_sheet.merge_cells(start_row=hd_row_max+1,start_column=ix,end_row=xls_max_rnum,end_column=ix)
		
		# font.bold
		add_bold = Font(name=dat_fnt,size=dat_fnt_sz,color='000000',bold=True,italic=False,vertAlign=None,underline='none',strike=False)
		for ix in col_bld_ix:
			for j in range(hd_row_max+1,xls_max_rnum):
				xl_cell.font = add_bold

		return 	   

	def tmplRead(self,
			tmpl_path: str = dst_path,
			sht_name: str = TMPL_CONFIG['TMPL_SHT_NAME'],
			bgn_col: str = TMPL_CONFIG['TMPL_BGN_COL'],
			end_col: str = TMPL_CONFIG['TMPL_END_COL'],
			header_row_max: int = TMPL_CONFIG['TMPL_HD_ROW'],
			footer_row_max: int = TMPL_CONFIG['TMPL_FT_ROW'],
			**kwagrs
		) -> dict:
		'''
		Read the stm template file, return a dict contains
		template header and foot data.
		
		Sample dict like:
			tmplData = {
				'header': (),
				'footer': ()
			}

		Reserved parameter options `kwargs` is to extend the template features 
		'''
		# Write your code here if additional features needed
		# if kwagrs is not None:
		# 	pass
		tmpl_data = {
			'header': None,
			'footer': None
		}
		header_sIX = bgn_col + '1'
		header_fIX = end_col + header_row_max
		footer_sIX = bgn_col + str(int(header_row_max) + 1)
		footer_fIX = end_col + footer_row_max

		# print([header_sIX, header_fIX, footer_sIX, footer_fIX])

		tmpl_wb = load_workbook(tmpl_path)
		tmpl_sheet = tmpl_wb[sht_name]
		
		tmpl_data['header'] = tmpl_sheet[header_sIX:header_fIX]
		tmpl_data['footer'] = tmpl_sheet[footer_sIX:footer_fIX]

		return tmpl_data

	def tmplWrite(self, 
			url: str = None,
			sa: str = None,
			output_dir: str = output_dir,
			filename: str = output_name,
			sht_name: str = TMPL_CONFIG['TMPL_SHT_NAME'],
			col_ix_max: int = 1024,   # to declare a huge list in case of any big table
			**kwagrs
		) -> None:
		'''templWrite doc strings'''
		assert (url is not None) \
		    or (sa is not None),'\n\t*** Warn: No input url/SA detected! ***\n'
	    
		sa = sa.upper() + '_'
		row_counter = 0

		# read template structure
		tmpl_data = self.tmplRead()
		wb = Workbook()
		ws = wb.active
		ws.title = sht_name

		# process headers
		# note that cell num is start from 1
		h_data = tmpl_data['header']
		row_num = len(h_data)
		col_num = len(h_data[0])
		for i in range(row_num):
			row_counter += 1
			for j in range(col_num):
				if h_data[i][j].value is None:
					continue
				else:
					_ = ws.cell(row=row_counter, column=j+1, value="{0}".format(h_data[i][j].value))

		# process body data
		# get tbl data from web
		from .WebCrawl import SiteTable

		tbl = SiteTable(url)
		tbl.requestInit()

		'''Following 4 variables are in TBL_ROW_IX list 
		   Test locals() and globals() method
		'''
		local = {}
		tbl_name = tbl.name
		tbl_desc = tbl.desc
		tbl_pks = tbl.pks
		tbl_strct = tbl.structure

		for row in tbl_strct['rows']:
			row_counter += 1
		
			'''Following 5 variables are in TBL_ROW_IX list'''
			src_col_name = row[0]
			src_col_type = row[1]
			src_col_desc = row[4]
			null_or_not = 'Y' if row[3] == 'X' else None
			tgt_dflt_value = row[2]

			if src_col_name.upper() in ['CREATION_DATE', 'LAST_MODIFIED_DATE']:
				row_counter -= 1
				continue

			tmp_row = [None for i in range(col_ix_max+1)]
			gdict = globals()
			ldict = locals()
			_ = [exec('tmp_row[{0}] = {1}'.format(key,value),gdict,ldict) for key, value in TBL_ROW_IX.items()]
			for j in range(col_num):
				if tmp_row[j] is None:
					continue
				else:
					_ = ws.cell(row=row_counter, column=j, value="{0}".format(tmp_row[j]))
		
		# process footer data
		f_data = tmpl_data['footer']
		row_num = len(f_data)
		col_num = len(f_data[0])
		row_counter = self.__footerHandler__(sa, f_data, row_counter, row_num, col_num, tbl_name, ws)
		
		self.__xlFormater__(
			xl_sheet=ws, 
			hd_row_max=int(TMPL_CONFIG['TMPL_HD_ROW']),
			xls_max_rnum=row_counter,
			xls_max_cnum=col_num
		)

		output_path = output_dir + 'STM_' + tbl_name + filename
		wb.save(output_path)
		print('\n\t*** Table: ' + tbl_name + ' has been processed! ***')
		print('\n\t*** Check it in: ' + output_path + ' ***\n')
		return

# if __name__ == '__main__':
# 	printf(MRG_CEL_IX)
# 	printf(isinstance(MRG_CEL_IX[1],int))